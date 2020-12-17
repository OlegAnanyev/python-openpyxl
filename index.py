import os
import time
from openpyxl import load_workbook


start = time.time() #засечём время работы скрипта

### ### ### ### ### ### ### ### ### ### ### ###
PATH_TO_MEGAFILE = 'L:\\python-openpyxl\\minifile.xlsx'
SHEET_NAME = 'TDSheet'
NEEDED_COLUMNS = [2, 6, 29] #какие столбцы оставляем в файле
PATH_TO_SAVEFILE = 'L:\\python-openpyxl\\result.xlsx'
### ### ### ### ### ### ### ### ### ### ### ###

wb = load_workbook(PATH_TO_MEGAFILE)
ws = wb[SHEET_NAME]

#########################################################################################################
#в цикле удаляем ненужные столбцы от последнего к первому, чтобы не нарушить последовательность
for col in range(ws.max_column, 0, -1):
    if col not in NEEDED_COLUMNS:
        ws.delete_cols(col)
#########################################################################################################

#########################################################################################################
#заменим все значения столбца на указанное (кроме заголовка)
COLUMN_TO_CHANGE = 1
NEW_VAL = 1234567890
for c in range (2, ws.max_row+1):
    ws.cell(row=c, column=COLUMN_TO_CHANGE).value = NEW_VAL
#########################################################################################################

#когда всё готово
wb.save(PATH_TO_SAVEFILE)
print ("New file is saved: ", PATH_TO_SAVEFILE)

elapsed = round (time.time() -start, 2)
print(elapsed, "seconds elapsed.")